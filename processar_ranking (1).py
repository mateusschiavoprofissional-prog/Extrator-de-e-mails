#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Processa um JSONL de mensagens enviadas do Gmail e gera:
- ranking_reencaminhamentos.xlsx
- ranking_reencaminhamentos_completo.csv
- ranking_reencaminhamentos_resultado.zip

Regra aplicada: 1 thread_id = 1 item no ranking; todos os envios dentro da thread contam igualmente.
"""

import csv
import json
import math
import zipfile
from collections import Counter, defaultdict
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter

INPUT_PATH = Path('/mnt/data/cache_mensagens_enviadas(2).jsonl')
XLSX_PATH = Path('/mnt/data/ranking_reencaminhamentos.xlsx')
CSV_PATH = Path('/mnt/data/ranking_reencaminhamentos_completo.csv')
ZIP_PATH = Path('/mnt/data/ranking_reencaminhamentos_resultado.zip')
SCRIPT_PATH = Path('/mnt/data/processar_ranking.py')

EXPECTED_COLUMNS = [
    'posição', 'thread_id', 'título completo', 'total_envios_thread', 'score_regra_total',
    'primeiro_envio', 'ultimo_envio', 'datas_horas_todos_envios', 'anos_todos_envios',
    'meses_todos_envios', 'message_ids'
]

WEIGHTS = {2026: 10, 2025: 8, 2024: 6, 2023: 4, 2022: 2}


def excel_safe_text(value):
    if value is None:
        return ''
    text = str(value)
    if text.startswith('='):
        text = "'" + text
    return text[:32767]


def parse_dt(obj):
    raw = obj.get('data_hora_envio') or obj.get('data_envio')
    if raw:
        for fmt in ('%d/%m/%Y %H:%M:%S', '%d/%m/%Y'):
            try:
                return datetime.strptime(str(raw), fmt)
            except ValueError:
                pass
    ts = obj.get('timestamp')
    if ts not in (None, ''):
        try:
            return datetime.fromtimestamp(float(ts))
        except Exception:
            return None
    return None


def get_year(obj, dt):
    try:
        return int(obj.get('ano'))
    except Exception:
        return dt.year if dt else None


def get_month(obj, dt):
    raw = obj.get('mes')
    if raw:
        return str(raw)
    return dt.strftime('%Y-%m') if dt else ''


def year_weight(year):
    return WEIGHTS.get(int(year), 1) if year is not None else 1


def format_dt(dt):
    return dt.strftime('%d/%m/%Y %H:%M:%S') if dt else ''


def process_jsonl(input_path):
    stats = {
        'total_lines': 0,
        'valid_messages': 0,
        'invalid_lines': 0,
        'duplicate_count': 0,
        'unique_messages': 0,
        'field_names': set(),
        'records': [],
    }
    seen_message_ids = set()
    with input_path.open('r', encoding='utf-8') as fh:
        for line in fh:
            stats['total_lines'] += 1
            try:
                obj = json.loads(line)
            except Exception:
                stats['invalid_lines'] += 1
                continue
            if not isinstance(obj, dict):
                stats['invalid_lines'] += 1
                continue
            stats['field_names'].update(obj.keys())
            message_id = obj.get('message_id')
            thread_id = obj.get('thread_id')
            if not message_id or not thread_id:
                stats['invalid_lines'] += 1
                continue
            stats['valid_messages'] += 1
            if message_id in seen_message_ids:
                stats['duplicate_count'] += 1
                continue
            seen_message_ids.add(message_id)
            dt = parse_dt(obj)
            year = get_year(obj, dt)
            month = get_month(obj, dt)
            stats['records'].append({
                'message_id': excel_safe_text(message_id),
                'thread_id': excel_safe_text(thread_id),
                'titulo_completo': excel_safe_text(obj.get('titulo_completo') or obj.get('titulo_original') or ''),
                'dt': dt,
                'dt_text': excel_safe_text(obj.get('data_hora_envio') or format_dt(dt)),
                'year': year,
                'month': month,
                'score': year_weight(year),
            })
    stats['field_names'] = sorted(stats['field_names'])
    stats['unique_messages'] = len(stats['records'])
    return stats


def build_ranking(records):
    grouped = defaultdict(list)
    for rec in records:
        grouped[rec['thread_id']].append(rec)

    ranking = []
    for thread_id, items in grouped.items():
        items_sorted = sorted(items, key=lambda x: (x['dt'] or datetime.min, x['message_id']))
        valid_dates = [r['dt'] for r in items_sorted if r['dt']]
        primeiro = min(valid_dates) if valid_dates else None
        ultimo = max(valid_dates) if valid_dates else None
        latest_item = max(items_sorted, key=lambda x: (x['dt'] or datetime.min, x['message_id']))
        ranking.append({
            'thread_id': thread_id,
            'título completo': latest_item['titulo_completo'] or items_sorted[0]['titulo_completo'],
            'total_envios_thread': len(items_sorted),
            'score_regra_total': sum(r['score'] for r in items_sorted),
            'primeiro_envio': primeiro,
            'ultimo_envio': ultimo,
            'datas_horas_todos_envios': excel_safe_text(' | '.join(r['dt_text'] for r in items_sorted)),
            'anos_todos_envios': excel_safe_text(' | '.join('' if r['year'] is None else str(r['year']) for r in items_sorted)),
            'meses_todos_envios': excel_safe_text(' | '.join(r['month'] for r in items_sorted)),
            'message_ids': excel_safe_text(' | '.join(r['message_id'] for r in items_sorted)),
        })

    ranking.sort(key=lambda r: (
        -r['score_regra_total'],
        -r['total_envios_thread'],
        -(r['ultimo_envio'].timestamp() if r['ultimo_envio'] else -math.inf),
        (r['título completo'] or '').casefold(),
    ))
    for pos, row in enumerate(ranking, 1):
        row['posição'] = pos
    return ranking


def build_monthly(records):
    counts = Counter()
    threads = defaultdict(set)
    for rec in records:
        if not rec['month']:
            continue
        if '-' in rec['month']:
            ano, mes = rec['month'].split('-', 1)
        else:
            ano, mes = '', rec['month']
        key = (ano, mes, rec['month'])
        counts[key] += 1
        threads[key].add(rec['thread_id'])
    return [[ano, mes, ym, counts[(ano, mes, ym)], len(threads[(ano, mes, ym)])] for ano, mes, ym in sorted(counts)]


def ranking_to_rows(ranking):
    rows = [EXPECTED_COLUMNS]
    for r in ranking:
        rows.append([
            r['posição'], r['thread_id'], r['título completo'], r['total_envios_thread'], r['score_regra_total'],
            r['primeiro_envio'], r['ultimo_envio'], r['datas_horas_todos_envios'], r['anos_todos_envios'],
            r['meses_todos_envios'], r['message_ids']
        ])
    return rows


def write_csv(ranking, csv_path):
    with csv_path.open('w', encoding='utf-8-sig', newline='') as fh:
        writer = csv.writer(fh, delimiter=';')
        writer.writerow(EXPECTED_COLUMNS)
        for r in ranking:
            writer.writerow([
                r['posição'], r['thread_id'], r['título completo'], r['total_envios_thread'], r['score_regra_total'],
                format_dt(r['primeiro_envio']), format_dt(r['ultimo_envio']), r['datas_horas_todos_envios'],
                r['anos_todos_envios'], r['meses_todos_envios'], r['message_ids']
            ])


def add_table(ws, name):
    if ws.max_row < 2 or ws.max_column < 1:
        return
    ref = f'A1:{get_column_letter(ws.max_column)}{ws.max_row}'
    table = Table(displayName=name, ref=ref)
    style = TableStyleInfo(name='TableStyleMedium2', showFirstColumn=False, showLastColumn=False,
                           showRowStripes=True, showColumnStripes=False)
    table.tableStyleInfo = style
    ws.add_table(table)


def style_sheet(ws, table_name=None, widths=None, highlight_headers=None):
    widths = widths or {}
    highlight_headers = set(highlight_headers or [])
    header_fill = PatternFill('solid', fgColor='1F4E78')
    key_fill = PatternFill('solid', fgColor='D9EAF7')
    white_font = Font(color='FFFFFF', bold=True)
    key_font = Font(color='000000', bold=True)
    border = Border(bottom=Side(style='thin', color='D9E2F3'))
    for cell in ws[1]:
        if cell.value in highlight_headers:
            cell.fill = key_fill
            cell.font = key_font
        else:
            cell.fill = header_fill
            cell.font = white_font
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = border
    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = ws.dimensions
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical='top', wrap_text=True)
            if cell.column in (6, 7) and isinstance(cell.value, datetime):
                cell.number_format = 'dd/mm/yyyy hh:mm:ss'
            if cell.column in (1, 4, 5):
                cell.number_format = '0'
    default_widths = {
        1: 10, 2: 22, 3: 52, 4: 18, 5: 18, 6: 22, 7: 22, 8: 65, 9: 28, 10: 35, 11: 65
    }
    default_widths.update(widths)
    for col_idx, width in default_widths.items():
        if col_idx <= ws.max_column:
            ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.row_dimensions[1].height = 28
    if table_name:
        add_table(ws, table_name)


def append_rows(ws, rows):
    for row in rows:
        ws.append(row)


def build_workbook(stats, ranking, monthly_rows):
    wb = Workbook()
    ws = wb.active
    ws.title = 'Resumo'
    mensal = wb.create_sheet('Resumo Mensal')
    top100 = wb.create_sheet('Top 100')
    completo = wb.create_sheet('Ranking Completo')

    records = stats['records']
    years = sorted({str(r['year']) for r in records if r['year'] is not None})
    months = sorted({r['month'] for r in records if r['month']})
    threads_unique = len({r['thread_id'] for r in records})
    one_send = sum(1 for r in ranking if r['total_envios_thread'] == 1)
    more_than_one = sum(1 for r in ranking if r['total_envios_thread'] > 1)
    all_dates = [r['dt'] for r in records if r['dt']]

    resumo_rows = [
        ['Indicador', 'Valor'],
        ['total de linhas lidas no JSONL', stats['total_lines']],
        ['total de mensagens válidas processadas', stats['valid_messages']],
        ['total de mensagens únicas', stats['unique_messages']],
        ['total de threads únicas', threads_unique],
        ['total de threads com 1 envio', one_send],
        ['total de threads com mais de 1 envio', more_than_one],
        ['total geral de envios', stats['unique_messages']],
        ['maior quantidade de envios em uma thread', max((r['total_envios_thread'] for r in ranking), default=0)],
        ['maior score encontrado', max((r['score_regra_total'] for r in ranking), default=0)],
        ['anos presentes na base', ', '.join(years)],
        ['meses presentes na base', ', '.join(months)],
        ['período inicial', min(all_dates) if all_dates else ''],
        ['período final', max(all_dates) if all_dates else ''],
        ['quantidade de duplicidades removidas', stats['duplicate_count']],
        ['quantidade de linhas inválidas', stats['invalid_lines']],
        ['campos encontrados no JSONL', ', '.join(stats['field_names'])],
    ]
    append_rows(ws, resumo_rows)
    style_sheet(ws, 'ResumoTable', widths={1: 45, 2: 95}, highlight_headers={'Indicador', 'Valor'})
    for cell in ('B13', 'B14'):
        ws[cell].number_format = 'dd/mm/yyyy hh:mm:ss'

    append_rows(mensal, [['ano', 'mês', 'ano_mes', 'quantidade_envios', 'quantidade_threads_distintas_no_mes']] + monthly_rows)
    style_sheet(mensal, 'ResumoMensalTable', widths={1: 12, 2: 12, 3: 14, 4: 20, 5: 35},
                highlight_headers={'quantidade_envios', 'quantidade_threads_distintas_no_mes'})

    append_rows(top100, ranking_to_rows(ranking[:100]))
    style_sheet(top100, 'Top100Table', highlight_headers={'posição', 'título completo', 'total_envios_thread', 'score_regra_total', 'primeiro_envio', 'ultimo_envio'})

    append_rows(completo, ranking_to_rows(ranking))
    style_sheet(completo, 'RankingCompletoTable', highlight_headers={'posição', 'título completo', 'total_envios_thread', 'score_regra_total', 'primeiro_envio', 'ultimo_envio'})

    return wb


def main():
    stats = process_jsonl(INPUT_PATH)
    ranking = build_ranking(stats['records'])
    monthly_rows = build_monthly(stats['records'])
    write_csv(ranking, CSV_PATH)
    wb = build_workbook(stats, ranking, monthly_rows)
    wb.save(XLSX_PATH)
    with zipfile.ZipFile(ZIP_PATH, 'w', compression=zipfile.ZIP_DEFLATED) as zf:
        zf.write(XLSX_PATH, XLSX_PATH.name)
        zf.write(CSV_PATH, CSV_PATH.name)
        zf.write(SCRIPT_PATH, SCRIPT_PATH.name)
    validation = {
        'total_linhas_lidas': stats['total_lines'],
        'mensagens_validas': stats['valid_messages'],
        'mensagens_unicas': stats['unique_messages'],
        'threads_unicas': len({r['thread_id'] for r in stats['records']}),
        'anos_encontrados': sorted({str(r['year']) for r in stats['records'] if r['year'] is not None}),
        'meses_encontrados': sorted({r['month'] for r in stats['records'] if r['month']}),
        'duplicidades_removidas': stats['duplicate_count'],
        'linhas_invalidas': stats['invalid_lines'],
        'maior_total_envios_thread': max((r['total_envios_thread'] for r in ranking), default=0),
        'maior_score': max((r['score_regra_total'] for r in ranking), default=0),
        'arquivos': [str(XLSX_PATH), str(CSV_PATH), str(ZIP_PATH), str(SCRIPT_PATH)]
    }
    print(json.dumps(validation, ensure_ascii=False, indent=2))


if __name__ == '__main__':
    main()
